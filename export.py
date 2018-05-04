from openpyxl import Workbook, load_workbook
from openpyxl.styles import PatternFill, Border, Side, Alignment, Protection, Font, numbers
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.worksheet.page import PageMargins
import charMenu
import cfgData
import json

charXlPath=cfgData.char_dir

def export_to_excel():
    p=charMenu.char_menu()
    menu_len=len(p)
    while True:
        s=int(raw_input("Select Character: "))
        if s >=1 and s<=menu_len:
            break
        else:
            print "Invalid Selection! Select a character from the list"

    # Open the file
    char_dict={}
    s-=1
    char = p[s]
    print char
    with open(cfgData.char_dir+"/"+char+"/"+char+".json","r") as cf:
        char_dict = json.load(cf)
    charXlFile=char_dict['Fullname']+".xlsx"
    with open(cfgData.cfg_dir+"/sttchart.csv") as f:
        statchart =f.read().splitlines()
    sc=[]

    for x in statchart:
        sc.append(x.split(","))

    # Loop through statistics to pull bonuses
    for x1 in sc:
        if int(x1[0]) == int(char_dict['st_stat']):
            stb,stdp,stpp=x1[1],x1[2],x1[3]
        if int(x1[0]) == int(char_dict['qu_stat']):
            qub,qudp,qupp=x1[1],x1[2],x1[3]
        if int(x1[0]) == int(char_dict['pr_stat']):
            prb,prdp,prpp=x1[1],x1[2],x1[3]
        if int(x1[0]) == int(char_dict['in_stat']):
            inb,indp,inpp=x1[1],x1[2],x1[3]
        if int(x1[0]) == int(char_dict['em_stat']):
            emb,emdp,empp=x1[1],x1[2],x1[3]
        if int(x1[0]) == int(char_dict['co_stat']):
            cob,codp,copp=x1[1],x1[2],x1[3]
        if int(x1[0]) == int(char_dict['ag_stat']):
            agb,agdp,agpp=x1[1],x1[2],x1[3]
        if int(x1[0]) == int(char_dict['sd_stat']):
            sdb,sddp,sdpp=x1[1],x1[2],x1[3]
        if int(x1[0]) == int(char_dict['me_stat']):
            meb,medp,mepp=x1[1],x1[2],x1[3]
        if int(x1[0]) == int(char_dict['re_stat']):
            reb,redp,repp=x1[1],x1[2],x1[3]
    tdp = float(codp) + float(agdp) + float(sddp) + float(medp) + float(redp)

    ###################
    # Lookup Race Bonus
    ###################
    with open(cfgData.cfg_dir+"/race.csv") as r:
        racechart =r.read().splitlines()
    rc=[]

    for x in racechart:
        rc.append(x.split(","))

    # Race Bonus
    for x2 in rc:
        if x2[0] == char_dict['race']:
            raceb=x2
    if char_dict['realm'] == "NULL":
        prpp,inpp,empp=0.0,0.0,0.0
        tpp=0.0
        realm="Non"
        realm_stat="N/A"
    if char_dict['realm'] == "PR":
        inpp,empp=0.0,0.0
        tpp=prpp
        realm="Mentalism"
        realm_stat="PR"
    if char_dict['realm'] == "IN":
        empp,prpp=0.0,0.0
        tpp=inpp
        realm="Channeling"
        realm_stat="IN"
    if char_dict['realm'] == "EM":
        inpp,prpp=0.0,0.0
        tpp=empp
        realm="Essence"
        realm_stat="EM"
    if char_dict['realm'] == "IP":
        empp=0.0
        tpp=(float(inpp)+float(prpp))/2
        realm="Channeling/Mentalism"
        realm_stat="IN/PR"
    if char_dict['realm'] == "PE":
        inpp=0.0
        tpp=(float(empp)+float(prpp))/2
        realm="Mentalism/Essence"
        realm_stat="PR/EM"
    if char_dict['realm'] == "IE":
        prpp=0.0
        tpp=(float(inpp)+float(empp))/2
        realm="Channeling/Essence"
        realm_stat="IN/EM"
    if char_dict['realm'] == "AR":
        tpp=(float(inpp)+float(prpp)+float(empp))/3
        realm="Arcane"
        realm_stat="IN/EM/PR"

    # Calculate Maneuver penalties
    if char_dict['armorType'] <= 4:
        current_mod = 0
    if char_dict['armorType'] >= 5 and char_dict['armorType'] <= 8:
        current_mod = int(char_dict['at_max_mod']) + int(char_dict['97'][14])
        if current_mod > int(char_dict['at_min_mod']):
            current_mod = char_dict['at_min_mod']
    if char_dict['armorType'] >= 9 and char_dict['armorType'] <= 12:
        current_mod = int(char_dict['at_max_mod']) + int(char_dict['98'][14])
        if current_mod > int(char_dict['at_min_mod']):
            current_mod = char_dict['at_min_mod']
    if char_dict['armorType'] >= 13 and char_dict['armorType'] <= 16:
        current_mod = int(char_dict['at_max_mod']) + int(char_dict['99'][14])
        if current_mod > int(char_dict['at_min_mod']):
            current_mod = char_dict['at_min_mod']
    if char_dict['armorType'] >= 17:
        current_mod = int(char_dict['at_max_mod']) + int(char_dict['100'][14])
        if current_mod > int(char_dict['at_min_mod']):
            current_mod = char_dict['at_min_mod']
    #print current_mod,":current_mod"

    ###### Create Excel file ######
    wb = Workbook()
    wb.create_sheet(index=1, title='Skills')
    # grab the active worksheet
    ws = wb.active
    ws.title = "Characters"
    ws = wb['Characters']
    ws1 = wb['Skills']
    ws.page_margins.bottom=0.5
    ws.page_margins.top=0.5
    ws1.page_margins.top=0.25
    ws1.page_margins.bottom=0.25
    ws1.page_margins.left=0.25
    ws1.page_margins.right=0.25

    # Data can be assigned directly to cells
    ws.column_dimensions['A'].width = 14.0
    ws.column_dimensions['B'].width = 8.25
    ws.column_dimensions['C'].width = 8.25
    ws.column_dimensions['E'].width = 10.85

    # Define colors
    gray = PatternFill("solid",fgColor="969696")
    lblue = PatternFill("solid",fgColor="CCFFFF")

    # Define number formatting

    # Define fonts
    textFont = Font(name='Arial',size=10)
    textBoldFont = Font(name='Arial',size=11, bold=True)
    textCalBoldFont = Font(name='Calibri',size=11, bold=True)
    headerFont = Font(name='Old English Text MT',size=14,bold=True)
    statFont = Font(name='Old English Text MT',size=18,bold=True)

    # Define Borders
    fullBorder = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
    tbBorder = Border(top=Side(style='thin'), bottom=Side(style='thin'))
    tblBorder = Border(top=Side(style='thin'), bottom=Side(style='thin'), left=Side(style='thin'))
    tbrBorder = Border(top=Side(style='thin'), bottom=Side(style='thin'), right=Side(style='thin'))
    rBorder = Border(right=Side(style='thin'))
    lBorder = Border(left=Side(style='thin'))
    tBorder = Border(top=Side(style='thin'))
    bBorder = Border(bottom=Side(style='thin'))
    tlrBorder = Border(top=Side(style='thin'), left=Side(style='thin'), right=Side(style='thin'))
    blrBorder = Border(left=Side(style='thin'), bottom=Side(style='thin'), right=Side(style='thin'))
    rlBorder = Border(left=Side(style='thin'), right=Side(style='thin'))
    rbBorder = Border(bottom=Side(style='thin'), right=Side(style='thin'))
    lbBorder = Border(bottom=Side(style='thin'), left=Side(style='thin'))
    tlBorder = Border(top=Side(style='thin'), left=Side(style='thin'))

    # Build Excel file
    ws.merge_cells('A1:I1')
    ws['A1'] = "Rolemaster Character Record"
    ws['A1'].font = headerFont
    ws['A1'].alignment = Alignment(horizontal='center',vertical='center')
    ws['A1'].border = tblBorder

    ws['B1'].border = tbBorder
    ws['C1'].border = tbBorder
    ws['D1'].border = tbBorder
    ws['E1'].border = tbBorder
    ws['F1'].border = tbBorder
    ws['G1'].border = tbBorder
    ws['H1'].border = tbBorder
    ws['I1'].border = tbrBorder

    ws['A2'] = "Name:"
    ws.merge_cells('B2:C2')
    ws['B2'] = char_dict['Fullname'].title()
    ws['E2'] = "Level:"
    ws['G2'] = "Race:"
    ws['F2'] = char_dict['lvl']
    ws['H2'] = char_dict['race']
    ws['F2'].alignment = Alignment(horizontal='left')
    ws['A2'].border = lBorder
    ws['I2'].border = rBorder
    ws['A2'].font = textFont
    ws['B2'].font = textFont
    ws['E2'].font = textFont
    ws['F2'].font = textFont
    ws['G2'].font = textFont
    ws['H2'].font = textFont

    ws['A3'] = "Profession:"
    ws.merge_cells('B3:C3')
    ws['B3'] = char_dict['proname']
    ws['E3'] = "Exp:"
    ws['F3'] = char_dict['exp']
    ws['G3'] = "Next Lvl:"
    ws['H3'] = char_dict['next_lvl']
    ws['F3'].number_format = "#,###"
    ws['H3'].number_format = "#,###"
    ws['E3'].alignment = Alignment(horizontal='right')
    ws['G3'].alignment = Alignment(horizontal='right')
    ws['A3'].border = lBorder
    ws['I3'].border = rBorder
    ws['A3'].font = textFont
    ws['B3'].font = textFont
    ws['E3'].font = textFont
    ws['F3'].font = textFont
    ws['G3'].font = textFont
    ws['H3'].font = textFont

    ws['A4'] = "Sex:"
    ws['B4'] = char_dict['gender']
    ws['C4'] = "Age:"
    ws['D4'] = char_dict['age']
    ws['F4'] = "Base Rate:"
    ws['C4'].alignment = Alignment(horizontal='right')
    ws['D4'].alignment = Alignment(horizontal='center')
    ws['F4'].alignment = Alignment(horizontal='right')
    ws['A4'].border = lBorder
    ws['I4'].border = rBorder
    ws['A4'].font = textFont
    ws['B4'].font = textFont
    ws['C4'].font = textFont
    ws['D4'].font = textFont
    ws['F4'].font = textFont
    ws['G4'].font = textFont

    ws['A5'] = "Height:"
    ws['C5'] = "Weight:"
    ws['F5'] = "Max Rate:"
    ws['B5'] = char_dict['height']
    ws['D5'] = char_dict['weight']
    ws['C5'].alignment = Alignment(horizontal='right')
    ws['D5'].alignment = Alignment(horizontal='center')
    ws['A5'].border = lBorder
    ws['I5'].border = rBorder
    ws['A5'].font = textFont
    ws['B5'].font = textFont
    ws['C5'].font = textFont
    ws['D5'].font = textFont
    ws['F5'].font = textFont
    ws['G5'].font = textFont

    ws['A6'] = "Hair:"
    ws['B6'] = char_dict['hair']
    ws['C6'] = "Eyes:"
    ws['D6'] = char_dict['eye']
    ws['C6'].alignment = Alignment(horizontal='right')
    ws['D6'].alignment = Alignment(horizontal='center')
    ws['A6'].border = lBorder
    ws['I6'].border = rBorder
    ws['A6'].font = textFont
    ws['B6'].font = textFont
    ws['C6'].font = textFont
    ws['D6'].font = textFont

    ws['A7'] = "Armor"
    ws.merge_cells('A7:D7')
    ws['A7'].alignment = Alignment(horizontal='center',vertical='center')
    ws['A7'].font = textBoldFont
    ws['E7'] = "Defensive Bonus"
    ws.merge_cells('E7:I7')
    ws['E7'].alignment = Alignment(horizontal='center',vertical='center')
    ws['E7'].font = textBoldFont
    ws['A7'].border = tblBorder
    ws['B7'].border = tbBorder
    ws['C7'].border = tbBorder
    ws['D7'].border = tbrBorder
    ws['E7'].border = tblBorder
    ws['F7'].border = tbBorder
    ws['G7'].border = tbBorder
    ws['H7'].border = tbBorder
    ws['I7'].border = tbrBorder

    ws['A8'] = "Armor Worn:"
    ws['B8'] = char_dict['armorTypeDesc']
    ws['E8'] = "Quickness:"
    ws['F8'] = int(qub)
    ws['F8'].alignment = Alignment(horizontal='center')
    ws['H8'] = "Total DB:"
    ws['A8'].font = textFont
    ws.merge_cells('B8:C8')
    ws['I8'] = '=IF(SUM(F8:F11)-F12<0,0,SUM(F8:F11)-F12)'
    ws['B8'].font = textFont
    ws['E8'].font = textFont
    ws['F8'].font = textFont
    ws['H8'].font = textFont
    ws['I8'].font = textFont
    ws['A8'].border = lBorder
    ws['E8'].border = lBorder
    ws['I8'].border = fullBorder
    ws['H8'].alignment = Alignment(horizontal='right')
    ws['I8'].alignment = Alignment(horizontal='center')

    ws['A9'] = "Armor Type:"
    ws['B9'] = char_dict['armorType']
    ws['E9'] = "Shield:"
    ws['F9'] = char_dict['shieldType']
    ws['H9'] = "vs Melee:"
    ws['I9'] = "=I8+{}".format(char_dict['shieldMeDB'])
    ws['A9'].font = textFont
    ws['B9'].font = textFont
    ws['E9'].font = textFont
    ws['F9'].font = textFont
    ws['H9'].font = textFont
    ws['I9'].font = textFont
    ws['A9'].border = lBorder
    ws['E9'].border = lBorder
    ws['I9'].border = fullBorder
    ws['B9'].alignment = Alignment(horizontal='center')
    ws['H9'].alignment = Alignment(horizontal='right')
    ws['I9'].alignment = Alignment(horizontal='center')

    ws['A10'] = "Shield:"
    ws['B10'] = char_dict['shieldType']
    ws['E10'] = "ADF:"
    ws['H10'] = "vs Missile:"
    ws['I10'] = '=I8+{}'.format(char_dict['shieldMiDB'])
    ws['A10'].font = textFont
    ws['B10'].font = textFont
    ws['E10'].font = textFont
    ws['F10'].font = textFont
    ws['H10'].font = textFont
    ws['I10'].font = textFont
    ws['A10'].border = lBorder
    ws['E10'].border = lBorder
    ws['I10'].border = fullBorder
    ws['H10'].alignment = Alignment(horizontal='right')
    ws['I10'].alignment = Alignment(horizontal='center')

    ws['A11'] = "Helm:"
    ws['B11'] = char_dict['helm']
    ws['E11'] = "Magical:"
    ws['H11'] = "vs Surprise:"
    ws['I11'] = '=I8-20'
    ws['A11'].font = textFont
    ws['B11'].font = textFont
    ws['E11'].font = textFont
    ws['F11'].font = textFont
    ws['H11'].font = textFont
    ws['I11'].font = textFont
    ws['A11'].border = lBorder
    ws['E11'].border = lBorder
    ws['I11'].border = fullBorder
    ws['H11'].alignment = Alignment(horizontal='right')
    ws['I11'].alignment = Alignment(horizontal='center')

    ws['A12'] = "Arm Greaves:"
    ws['B12'] = char_dict['armGrea']
    ws['E12'] = "Armor Pen:"
    ws['F12'] = int(char_dict['at_qu_pen'])
    ws['H12'] = "vs Magic:"
    ws['I12'] = "=I8+{}+{}".format(char_dict['shieldMaDB'],char_dict['helmMaDB'])
    ws['A12'].font = textFont
    ws['B12'].font = textFont
    ws['E12'].font = textFont
    ws['F12'].font = textFont
    ws['H12'].font = textFont
    ws['I12'].font = textFont
    ws['A12'].border = lBorder
    ws['E12'].border = lBorder
    ws['I12'].border = fullBorder
    ws['H12'].alignment = Alignment(horizontal='right')
    ws['I12'].alignment = Alignment(horizontal='center')
    ws['F12'].alignment = Alignment(horizontal='center')

    ws['A13'] = "Leg Greaves:"
    ws['B13'] = char_dict['legGrea']
    ws['E13'] = "Hits"
    ws['A13'].font = textFont
    ws['B13'].font = textFont
    ws['E13'].font = textBoldFont
    ws.merge_cells('E13:H13')
    ws['E13'].alignment = Alignment(horizontal='center',vertical='center')
    ws['A13'].border = lBorder
    ws['E13'].border = tblBorder
    ws['F13'].border = tbBorder
    ws['G13'].border = tbBorder
    ws['H13'].border = tbrBorder
    ws['I13'].border = rBorder

    ws['A14'] = "Maneuver Modifiers"
    ws.merge_cells('A14:C14')
    ws['A14'].font = textBoldFont
    ws['A14'].alignment = Alignment(horizontal='center',vertical='center')
    ws['E14'] = "Total Hits:"
    ws['G14'] = "Hit Die:"
    ws['E14'].font = textFont
    ws['F14'].font = textFont
    ws['G14'].font = textFont
    ws['H14'].font = textFont
    ws['A14'].border = lBorder
    ws['E14'].border = lBorder
    ws['H14'].border = rBorder
    ws['I14'].border = rBorder
    ws['H14'] = char_dict['HitDie']
    ws['H14'].alignment = Alignment(horizontal='center')

    ws['A15'] = "Minimum:"
    ws['B15'] = char_dict['at_min_mod']
    ws['B15'].alignment = Alignment(horizontal='center')
    ws['C15'] = "Current"
    ws['E15'] = "Base Hits:"
    ws['F15'] = char_dict['hp_base']
    ws['F15'].alignment = Alignment(horizontal='center')
    ws['A15'].font = textFont
    ws['B15'].font = textFont
    ws['C15'].font = textFont
    ws['E15'].font = textFont
    ws['F15'].font = textFont
    ws['A15'].border = lBorder
    ws['E15'].border = lBorder
    ws['H15'].border = rBorder
    ws['I15'].border = rBorder

    ws['A16'] = "Maximum:"
    ws['B16'] = char_dict['at_max_mod']
    ws['B16'].alignment = Alignment(horizontal='center')
    ws['C16'] = current_mod
    ws['C16'].alignment = Alignment(horizontal='center')
    ws['C16'].font = textBoldFont
    ws['A16'].font = textFont
    ws['B16'].font = textFont
    ws['A16'].border = lBorder
    ws['E16'].border = lBorder
    ws['H16'].border = rBorder
    ws['I16'].border = rBorder

    ws['A17'] = "Languages"
    ws['A17'].font = textBoldFont
    ws['A17'].alignment = Alignment(horizontal='center')
    ws['A17'].border = tlBorder
    ws.merge_cells('A17:A18')
    ws['B17'].border = tBorder
    ws.merge_cells('C17:D17')
    ws['C17'] = "-- Rank --"
    ws['C17'].font = textFont
    ws['C17'].border = tlrBorder
    ws['C17'].alignment = Alignment(horizontal='center')
    ws['D17'].border = tBorder
    ws['E17'] = "Magical Statistics"
    ws.merge_cells('E17:H17')
    ws['E17'].font = textBoldFont
    ws['E17'].alignment = Alignment(horizontal='center',vertical='center')
    ws['E17'].border = tblBorder
    ws['F17'].border = tbBorder
    ws['G17'].border = tbBorder
    ws['H17'].border = tbrBorder
    ws['I17'].border = rBorder

    ws['C18'] = "Spoken"
    ws['D18'] = "Written"
    ws['E18'] = "Realm:"
    ws['A18'].border = lbBorder
    ws['B18'].border = bBorder
    ws['C18'].border = fullBorder
    ws['D18'].border = fullBorder
    ws['E18'].border = lBorder
    ws['F18'] = char_dict['realm']
    ws['H18'].border = rBorder
    ws['I18'].border = rBorder
    ws['C18'].font = textFont
    ws['D18'].font = textFont
    ws['E18'].font = textFont
    ws['F18'].font = textFont
    ws.merge_cells('F18:G18')
    ws['C18'].alignment = Alignment(horizontal='center')
    ws['D18'].alignment = Alignment(horizontal='center')

    if char_dict.has_key("lang1") == False:
        char_dict['lang1'] = ["","",""]
    lang1 = char_dict['lang1'][0]
    lang1s = char_dict['lang1'][1]
    lang1w = char_dict['lang1'][2]
    ws.merge_cells('A19:B19')
    ws['A19'].font = textFont
    ws['A19'] = lang1
    ws['C19'] = lang1s
    ws['D19'] = lang1w
    ws['F19'] = char_dict['mrealm']
    ws['C19'].alignment = Alignment(horizontal='center')
    ws['D19'].alignment = Alignment(horizontal='center')
    ws['E19'] = "Stat Bonus:"
    ws['A19'].border = lBorder
    ws['C19'].border = fullBorder
    ws['D19'].border = fullBorder
    ws['E19'].border = lBorder
    ws['H19'].border = rBorder
    ws['I19'].border = rBorder
    ws['C19'].font = textFont
    ws['D19'].font = textFont
    ws['E19'].font = textFont
    ws['F19'].font = textFont

    if char_dict.has_key("lang2") == False:
        char_dict['lang2'] = ["","",""]
    lang2 = char_dict['lang2'][0]
    lang2s = char_dict['lang2'][1]
    lang2w = char_dict['lang2'][2]
    ws.merge_cells('A20:B20')
    ws['A20'] = lang2
    ws['C20'] = lang2s
    ws['D20'] = lang2w
    ws['C20'].alignment = Alignment(horizontal='center')
    ws['D20'].alignment = Alignment(horizontal='center')
    ws['A20'].font = textFont
    ws['E20'] = "Lvl Bonus:"
    ws['A20'].border = lBorder
    ws['C20'].border = fullBorder
    ws['D20'].border = fullBorder
    ws['E20'].border = lBorder
    ws['H20'].border = rBorder
    ws['I20'].border = rBorder
    ws['C20'].font = textFont
    ws['D20'].font = textFont
    ws['E20'].font = textFont
    ws['F20'].font = textFont

    if char_dict.has_key("lang3") == False:
        char_dict['lang3'] = ["","",""]
    lang3 = char_dict['lang3'][0]
    lang3s = char_dict['lang3'][1]
    lang3w = char_dict['lang3'][2]
    ws.merge_cells('A21:B21')
    ws['A21'] = lang3
    ws['C21'] = lang3s
    ws['D21'] = lang3w
    ws['C21'].alignment = Alignment(horizontal='center')
    ws['D21'].alignment = Alignment(horizontal='center')
    ws['A21'].font = textFont
    ws['E21'] = "Resistance Roll Modifiers"
    ws.merge_cells('E21:I21')
    ws['E21'].font = textBoldFont
    ws['E21'].alignment = Alignment(horizontal='center',vertical='center')
    ws['A21'].border = lBorder
    ws['C21'].border = fullBorder
    ws['D21'].border = fullBorder
    ws['E21'].border = tblBorder
    ws['F21'].border = tbBorder
    ws['G21'].border = tbBorder
    ws['H21'].border = tbBorder
    ws['I21'].border = tbrBorder

    if char_dict.has_key('lang4') == False:
        char_dict['lang4'] = ["","",""]
    lang4 = char_dict['lang4'][0]
    lang4s = char_dict['lang4'][1]
    lang4w = char_dict['lang4'][2]
    ws.merge_cells('A22:B22')
    ws['A22'] = lang4
    ws['C22'] = lang4s
    ws['D22'] = lang4w
    ws['F22'] = char_dict['Essmod']
    ws['I22'] = char_dict['sdtb']
    ws['C22'].alignment = Alignment(horizontal='center')
    ws['D22'].alignment = Alignment(horizontal='center')
    ws['F22'].alignment = Alignment(horizontal='center')
    ws['I22'].alignment = Alignment(horizontal='center')
    ws['A22'].font = textFont
    ws['C22'].font = textFont
    ws['D22'].font = textFont
    ws['E22'].font = textFont
    ws['F22'].font = textFont
    ws['H22'].font = textFont
    ws['I22'].font = textFont
    ws['E22'] = "Essence:"
    ws['H22'] = "Fear:"
    ws['E22'].font = textFont
    ws['E22'].alignment = Alignment(horizontal='right')
    ws['H22'].alignment = Alignment(horizontal='right')
    ws['A22'].border = lBorder
    ws['C22'].border = fullBorder
    ws['D22'].border = fullBorder
    ws['I22'].border = rBorder

    if char_dict.has_key('lang5') == False:
        char_dict['lang5'] = ["","",""]
    lang5 = char_dict['lang5'][0]
    lang5s = char_dict['lang5'][1]
    lang5w = char_dict['lang5'][2]
    ws.merge_cells('A23:B23')
    ws['A23'] = lang5
    ws['C23'] = lang5s
    ws['D23'] = lang5w
    ws['F23'] = char_dict['Mentmod']
    ws['I23'] = char_dict['Poimod']
    ws['C23'].alignment = Alignment(horizontal='center')
    ws['D23'].alignment = Alignment(horizontal='center')
    ws['F23'].alignment = Alignment(horizontal='center')
    ws['I23'].alignment = Alignment(horizontal='center')
    ws['A23'].font = textFont
    ws['C23'].font = textFont
    ws['D23'].font = textFont
    ws['E23'].font = textFont
    ws['F23'].font = textFont
    ws['H23'].font = textFont
    ws['I23'].font = textFont
    ws['E23'] = "Mentalism:"
    ws['H23'] = "Poison:"
    ws['E23'].font = textFont
    ws['E23'].alignment = Alignment(horizontal='right')
    ws['H23'].alignment = Alignment(horizontal='right')
    ws['A23'].border = lBorder
    ws['C23'].border = fullBorder
    ws['D23'].border = fullBorder
    ws['I23'].border = rBorder

    if char_dict.has_key('lang6') == False:
        char_dict['lang6'] = ["","",""]
    lang6 = char_dict['lang6'][0]
    lang6s = char_dict['lang6'][1]
    lang6w = char_dict['lang6'][2]
    ws.merge_cells('A24:B24')
    ws['A24'] = lang6
    ws['C24'] = lang6s
    ws['D24'] = lang6w
    ws['F24'] = char_dict['Chanmod']
    ws['I24'] = char_dict['Dismod']
    ws['C24'].alignment = Alignment(horizontal='center')
    ws['D24'].alignment = Alignment(horizontal='center')
    ws['F24'].alignment = Alignment(horizontal='center')
    ws['I24'].alignment = Alignment(horizontal='center')
    ws['A24'].font = textFont
    ws['C24'].font = textFont
    ws['D24'].font = textFont
    ws['E24'].font = textFont
    ws['F24'].font = textFont
    ws['H24'].font = textFont
    ws['I24'].font = textFont
    ws['E24'] = "Channeling:"
    ws['H24'] = "Disease:"
    ws['E24'].font = textFont
    ws['E24'].alignment = Alignment(horizontal='right')
    ws['H24'].alignment = Alignment(horizontal='right')
    ws['A24'].border = lBorder
    ws['C24'].border = fullBorder
    ws['D24'].border = fullBorder
    ws['I24'].border = rBorder

    ws['A25'] = "Background Options / Miscellaneous Notes"
    ws.merge_cells('A25:I25')
    ws['A25'].alignment = Alignment(horizontal='center',vertical='center')
    ws['A25'].font = textBoldFont
    ws['A25'].border = tblBorder
    ws['B25'].border = tbBorder
    ws['C25'].border = tbBorder
    ws['D25'].border = tbBorder
    ws['E25'].border = tbBorder
    ws['F25'].border = tbBorder
    ws['G25'].border = tbBorder
    ws['H25'].border = tbBorder
    ws['I25'].border = tbrBorder

    ws.merge_cells('A26:I26')
    ws['A26'].border = lBorder
    ws['I26'].border = rBorder

    ws.merge_cells('A27:I27')
    ws['A27'].border = lBorder
    ws['I27'].border = rBorder

    ws.merge_cells('A28:I28')
    ws['A28'].border = lBorder
    ws['I28'].border = rBorder

    ws.merge_cells('A29:I29')
    ws['A29'].border = lBorder
    ws['I29'].border = rBorder

    ws.merge_cells('A30:I30')
    ws['A30'].border = lBorder
    ws['I30'].border = rBorder

    ws.merge_cells('A31:I31')
    ws['A31'].border = lBorder
    ws['I31'].border = rBorder

    ws.merge_cells('A32:I32')
    ws['A32'].border = lBorder
    ws['I32'].border = rBorder

    ws.merge_cells('A33:I33')
    ws['A33'].border = lbBorder
    ws['B33'].border = bBorder
    ws['C33'].border = bBorder
    ws['D33'].border = bBorder
    ws['E33'].border = bBorder
    ws['F33'].border = bBorder
    ws['G33'].border = bBorder
    ws['H33'].border = bBorder
    ws['I33'].border = rbBorder

    ws.merge_cells('A34:A35')
    ws['A34'].font = statFont
    ws['A34'] = "Stats"
    ws.merge_cells('B34:B35')
    ws.merge_cells('C34:C35')
    ws.merge_cells('D34:D35')
    ws.merge_cells('E34:G34')
    ws.merge_cells('H34:H35')
    ws.merge_cells('I34:I35')
    ws['B34'].font = textBoldFont
    ws['C34'].font = textBoldFont
    ws['D34'].font = textBoldFont
    ws['E34'].font = textBoldFont
    ws['E35'].font = textBoldFont
    ws['F35'].font = textBoldFont
    ws['G35'].font = textBoldFont
    ws['H34'].font = textBoldFont
    ws['I34'].font = textBoldFont
    ws['B34'].alignment = Alignment(horizontal='center',vertical='center')
    ws['C34'].alignment = Alignment(horizontal='center',vertical='center')
    ws['D34'].alignment = Alignment(horizontal='center',vertical='center')
    ws['E34'].alignment = Alignment(horizontal='center',vertical='center')
    ws['E35'].alignment = Alignment(horizontal='center')
    ws['F35'].alignment = Alignment(horizontal='center')
    ws['G35'].alignment = Alignment(horizontal='center')
    ws['H34'].alignment = Alignment(horizontal='center',vertical='center',wrapText=True)
    ws['I34'].alignment = Alignment(horizontal='center',vertical='center',wrapText=True)
    ws['A34'].border = lBorder
    ws['A35'].border = lBorder
    ws['B34'].border = lBorder
    ws['B35'].border = lBorder
    ws['C34'].border = lBorder
    ws['C35'].border = lBorder
    ws['D34'].border = lBorder
    ws['D35'].border = lBorder
    ws['E34'].border = lBorder
    ws['E35'].border = fullBorder
    ws['F35'].border = fullBorder
    ws['G35'].border = fullBorder
    ws['H34'].border = lBorder
    ws['H35'].border = lBorder
    ws['I34'].border = rlBorder
    ws['I35'].border = rlBorder
    ws['B34'] = "Cur"
    ws['C34'] = "Pot"
    ws['D34'] = "Total"
    ws['E34'] = "-- BONUSES --"
    ws['E35'] = "Stat"
    ws['F35'] = "Race"
    ws['G35'] = "Misc"
    ws['H34'] = "Dev. Pts"
    ws['I34'] = "Power Pts"

    ws['A36'] = "Strength"
    ws['A36'].font = textFont
    ws['A36'].border = fullBorder
    ws['B36'] = char_dict['st_stat']
    ws['B36'].alignment = Alignment(horizontal='center')
    ws['B36'].font = textFont
    ws['B36'].border = fullBorder
    ws['C36'] = char_dict['st_pot']
    ws['C36'].alignment = Alignment(horizontal='center')
    ws['C36'].font = textFont
    ws['C36'].border = fullBorder
    ws['D36'] = char_dict['sttb']
    ws['D36'].alignment = Alignment(horizontal='center')
    ws['D36'].font = textFont
    ws['D36'].border = fullBorder
    ws['E36'] = int(stb)
    ws['E36'].alignment = Alignment(horizontal='center')
    ws['E36'].font = textFont
    ws['E36'].border = fullBorder
    ws['F36'] = int(raceb[1])
    ws['F36'].alignment = Alignment(horizontal='center')
    ws['F36'].font = textFont
    ws['F36'].border = fullBorder
    ws['G36'] = char_dict['stmb']
    ws['G36'].alignment = Alignment(horizontal='center')
    ws['G36'].font = textFont
    ws['G36'].border = fullBorder
    ws['H36'].font = textFont
    ws['H36'].border = fullBorder
    ws['I36'].font = textFont
    ws['I36'].border = fullBorder
    ws['H36'].fill = gray
    ws['I36'].fill = gray

    ws['A37'] = "Quickness"
    ws['A37'].font = textFont
    ws['A37'].border = fullBorder
    ws['B37'] = char_dict['qu_stat']
    ws['B37'].alignment = Alignment(horizontal='center')
    ws['B37'].font = textFont
    ws['B37'].border = fullBorder
    ws['C37'] = char_dict['qu_pot']
    ws['C37'].alignment = Alignment(horizontal='center')
    ws['C37'].font = textFont
    ws['C37'].border = fullBorder
    ws['D37'] = char_dict['qutb']
    ws['D37'].alignment = Alignment(horizontal='center')
    ws['D37'].font = textFont
    ws['D37'].border = fullBorder
    ws['E37'] = int(qub)
    ws['E37'].alignment = Alignment(horizontal='center')
    ws['E37'].font = textFont
    ws['E37'].border = fullBorder
    ws['F37'] = int(raceb[2])
    ws['F37'].alignment = Alignment(horizontal='center')
    ws['F37'].font = textFont
    ws['F37'].border = fullBorder
    ws['G37'] = char_dict['qumb']
    ws['G37'].alignment = Alignment(horizontal='center')
    ws['G37'].font = textFont
    ws['G37'].border = fullBorder
    ws['H37'].font = textFont
    ws['H37'].border = fullBorder
    ws['I37'].font = textFont
    ws['I37'].border = fullBorder
    ws['H37'].fill = gray
    ws['I37'].fill = gray

    ws['A38'] = "Presence"
    ws['A38'].font = textFont
    ws['A38'].border = fullBorder
    ws['B38'] = char_dict['pr_stat']
    ws['B38'].alignment = Alignment(horizontal='center')
    ws['B38'].font = textFont
    ws['B38'].border = fullBorder
    ws['C38'] = char_dict['pr_pot']
    ws['C38'].alignment = Alignment(horizontal='center')
    ws['C38'].font = textFont
    ws['C38'].border = fullBorder
    ws['D38'] = char_dict['prtb']
    ws['D38'].alignment = Alignment(horizontal='center')
    ws['D38'].font = textFont
    ws['D38'].border = fullBorder
    ws['E38'] = int(prb)
    ws['E38'].alignment = Alignment(horizontal='center')
    ws['E38'].font = textFont
    ws['E38'].border = fullBorder
    ws['F38'] = int(raceb[3])
    ws['F38'].alignment = Alignment(horizontal='center')
    ws['F38'].font = textFont
    ws['F38'].border = fullBorder
    ws['G38'] = char_dict['prmb']
    ws['G38'].alignment = Alignment(horizontal='center')
    ws['G38'].font = textFont
    ws['G38'].border = fullBorder
    ws['H38'].font = textFont
    ws['H38'].border = fullBorder
    ws['I38'] = prpp
    ws['I38'].alignment = Alignment(horizontal='center')
    ws['I38'].font = textFont
    ws['I38'].border = fullBorder
    ws['I38'].number_format = '#,##0.0'
    ws['H38'].fill = gray

    ws['A39'] = "Intuition"
    ws['A39'].font = textFont
    ws['A39'].border = fullBorder
    ws['B39'] = char_dict['in_stat']
    ws['B39'].alignment = Alignment(horizontal='center')
    ws['B39'].font = textFont
    ws['B39'].border = fullBorder
    ws['C39'] = char_dict['in_pot']
    ws['C39'].alignment = Alignment(horizontal='center')
    ws['C39'].font = textFont
    ws['C39'].border = fullBorder
    ws['D39'] = char_dict['intb']
    ws['D39'].alignment = Alignment(horizontal='center')
    ws['D39'].font = textFont
    ws['D39'].border = fullBorder
    ws['E39'] = int(inb)
    ws['E39'].alignment = Alignment(horizontal='center')
    ws['E39'].font = textFont
    ws['E39'].border = fullBorder
    ws['F39'] = int(raceb[4])
    ws['F39'].alignment = Alignment(horizontal='center')
    ws['F39'].font = textFont
    ws['F39'].border = fullBorder
    ws['G39'] = char_dict['inmb']
    ws['G39'].alignment = Alignment(horizontal='center')
    ws['G39'].font = textFont
    ws['G39'].border = fullBorder
    ws['H39'].font = textFont
    ws['H39'].border = fullBorder
    ws['I39'] = inpp
    ws['I39'].alignment = Alignment(horizontal='center')
    ws['I39'].font = textFont
    ws['I39'].border = fullBorder
    ws['I39'].number_format = '#,##0.0'
    ws['H39'].fill = gray

    ws['A40'] = "Empathy"
    ws['A40'].font = textFont
    ws['A40'].border = fullBorder
    ws['B40'] = char_dict['em_stat']
    ws['B40'].alignment = Alignment(horizontal='center')
    ws['B40'].font = textFont
    ws['B40'].border = fullBorder
    ws['C40'] = char_dict['em_pot']
    ws['C40'].alignment = Alignment(horizontal='center')
    ws['C40'].font = textFont
    ws['C40'].border = fullBorder
    ws['D40'] = char_dict['emtb']
    ws['D40'].alignment = Alignment(horizontal='center')
    ws['D40'].font = textFont
    ws['D40'].border = fullBorder
    ws['E40'] = int(emb)
    ws['E40'].alignment = Alignment(horizontal='center')
    ws['E40'].font = textFont
    ws['E40'].border = fullBorder
    ws['F40'] = int(raceb[5])
    ws['F40'].alignment = Alignment(horizontal='center')
    ws['F40'].font = textFont
    ws['F40'].border = fullBorder
    ws['G40'] = char_dict['emmb']
    ws['G40'].alignment = Alignment(horizontal='center')
    ws['G40'].font = textFont
    ws['G40'].border = fullBorder
    ws['H40'].font = textFont
    ws['H40'].border = fullBorder
    ws['I40'] = empp
    ws['I40'].alignment = Alignment(horizontal='center')
    ws['I40'].font = textFont
    ws['I40'].border = fullBorder
    ws['I40'].number_format = '#,##0.0'
    ws['H40'].fill = gray

    ws['A41'] = "Constitution"
    ws['A41'].font = textFont
    ws['A41'].border = fullBorder
    ws['B41'] = char_dict['co_stat']
    ws['B41'].alignment = Alignment(horizontal='center')
    ws['B41'].font = textFont
    ws['B41'].border = fullBorder
    ws['C41'] = char_dict['co_pot']
    ws['C41'].alignment = Alignment(horizontal='center')
    ws['C41'].font = textFont
    ws['C41'].border = fullBorder
    ws['D41'] = char_dict['cotb']
    ws['D41'].alignment = Alignment(horizontal='center')
    ws['D41'].font = textFont
    ws['D41'].border = fullBorder
    ws['E41'] = int(cob)
    ws['E41'].alignment = Alignment(horizontal='center')
    ws['E41'].font = textFont
    ws['E41'].border = fullBorder
    ws['F41'] = int(raceb[6])
    ws['F41'].alignment = Alignment(horizontal='center')
    ws['F41'].font = textFont
    ws['F41'].border = fullBorder
    ws['G41'] = char_dict['comb']
    ws['G41'].alignment = Alignment(horizontal='center')
    ws['G41'].font = textFont
    ws['G41'].border = fullBorder
    ws['H41'] = float(codp)
    ws['H41'].alignment = Alignment(horizontal='center')
    ws['H41'].font = textFont
    ws['H41'].border = fullBorder
    ws['H41'].number_format = '#,##0.0'
    ws['I41'].font = textFont
    ws['I41'].border = fullBorder
    ws['I41'].fill = gray

    ws['A42'] = "Agility"
    ws['A42'].font = textFont
    ws['A42'].border = fullBorder
    ws['B42'] = char_dict['ag_stat']
    ws['B42'].alignment = Alignment(horizontal='center')
    ws['B42'].font = textFont
    ws['B42'].border = fullBorder
    ws['C42'] = char_dict['ag_pot']
    ws['C42'].alignment = Alignment(horizontal='center')
    ws['C42'].font = textFont
    ws['C42'].border = fullBorder
    ws['D42'] = char_dict['agtb']
    ws['D42'].alignment = Alignment(horizontal='center')
    ws['D42'].font = textFont
    ws['D42'].border = fullBorder
    ws['E42'] = int(agb)
    ws['E42'].alignment = Alignment(horizontal='center')
    ws['E42'].font = textFont
    ws['E42'].border = fullBorder
    ws['F42'] = int(raceb[7])
    ws['F42'].alignment = Alignment(horizontal='center')
    ws['F42'].font = textFont
    ws['F42'].border = fullBorder
    ws['G42'] = char_dict['agmb']
    ws['G42'].alignment = Alignment(horizontal='center')
    ws['G42'].font = textFont
    ws['G42'].border = fullBorder
    ws['H42'] = float(agdp)
    ws['H42'].alignment = Alignment(horizontal='center')
    ws['H42'].font = textFont
    ws['H42'].border = fullBorder
    ws['H42'].number_format = '#,##0.0'
    ws['I42'].font = textFont
    ws['I42'].border = fullBorder
    ws['I42'].fill = gray

    ws['A43'] = "Self Discipline"
    ws['A43'].font = textFont
    ws['A43'].border = fullBorder
    ws['B43'] = char_dict['sd_stat']
    ws['B43'].alignment = Alignment(horizontal='center')
    ws['B43'].font = textFont
    ws['B43'].border = fullBorder
    ws['C43'] = char_dict['sd_pot']
    ws['C43'].alignment = Alignment(horizontal='center')
    ws['C43'].font = textFont
    ws['C43'].border = fullBorder
    ws['D43'] = char_dict['sdtb']
    ws['D43'].alignment = Alignment(horizontal='center')
    ws['D43'].font = textFont
    ws['D43'].border = fullBorder
    ws['E43'] = int(sdb)
    ws['E43'].alignment = Alignment(horizontal='center')
    ws['E43'].font = textFont
    ws['E43'].border = fullBorder
    ws['F43'] = int(raceb[8])
    ws['F43'].alignment = Alignment(horizontal='center')
    ws['F43'].font = textFont
    ws['F43'].border = fullBorder
    ws['G43'] = char_dict['sdmb']
    ws['G43'].alignment = Alignment(horizontal='center')
    ws['G43'].font = textFont
    ws['G43'].border = fullBorder
    ws['H43'] = float(sddp)
    ws['H43'].alignment = Alignment(horizontal='center')
    ws['H43'].font = textFont
    ws['H43'].border = fullBorder
    ws['H43'].number_format = '#,##0.0'
    ws['I43'].font = textFont
    ws['I43'].border = fullBorder
    ws['I43'].fill = gray

    ws['A44'] = "Memory"
    ws['A44'].font = textFont
    ws['A44'].border = fullBorder
    ws['B44'] = char_dict['me_stat']
    ws['B44'].alignment = Alignment(horizontal='center')
    ws['B44'].font = textFont
    ws['B44'].border = fullBorder
    ws['C44'] = char_dict['me_pot']
    ws['C44'].alignment = Alignment(horizontal='center')
    ws['C44'].font = textFont
    ws['C44'].border = fullBorder
    ws['D44'] = char_dict['metb']
    ws['D44'].alignment = Alignment(horizontal='center')
    ws['D44'].font = textFont
    ws['D44'].border = fullBorder
    ws['E44'] = int(meb)
    ws['E44'].alignment = Alignment(horizontal='center')
    ws['E44'].font = textFont
    ws['E44'].border = fullBorder
    ws['F44'] = int(raceb[9])
    ws['F44'].alignment = Alignment(horizontal='center')
    ws['F44'].font = textFont
    ws['F44'].border = fullBorder
    ws['G44'] = char_dict['memb']
    ws['G44'].alignment = Alignment(horizontal='center')
    ws['G44'].font = textFont
    ws['G44'].border = fullBorder
    ws['H44'] = float(medp)
    ws['H44'].alignment = Alignment(horizontal='center')
    ws['H44'].font = textFont
    ws['H44'].border = fullBorder
    ws['H44'].number_format = '#,##0.0'
    ws['I44'].font = textFont
    ws['I44'].border = fullBorder
    ws['I44'].fill = gray

    ws['A45'] = "Reasoning"
    ws['A45'].font = textFont
    ws['A45'].border = fullBorder
    ws['B45'] = char_dict['re_stat']
    ws['B45'].alignment = Alignment(horizontal='center')
    ws['B45'].font = textFont
    ws['B45'].border = fullBorder
    ws['C45'] = char_dict['re_pot']
    ws['C45'].alignment = Alignment(horizontal='center')
    ws['C45'].font = textFont
    ws['C45'].border = fullBorder
    ws['D45'] = char_dict['retb']
    ws['D45'].alignment = Alignment(horizontal='center')
    ws['D45'].font = textFont
    ws['D45'].border = fullBorder
    ws['E45'] = int(reb)
    ws['E45'].alignment = Alignment(horizontal='center')
    ws['E45'].font = textFont
    ws['E45'].border = fullBorder
    ws['F45'] = int(raceb[10])
    ws['F45'].alignment = Alignment(horizontal='center')
    ws['F45'].font = textFont
    ws['F45'].border = fullBorder
    ws['G45'] = char_dict['remb']
    ws['G45'].alignment = Alignment(horizontal='center')
    ws['G45'].font = textFont
    ws['G45'].border = fullBorder
    ws['H45'] = float(redp)
    ws['H45'].alignment = Alignment(horizontal='center')
    ws['H45'].font = textFont
    ws['H45'].border = fullBorder
    ws['H45'].number_format = '#,##0.0'
    ws['I45'].font = textFont
    ws['I45'].border = fullBorder
    ws['I45'].fill = gray

    ws['H46'].font = textFont
    ws['H46'].border = fullBorder
    ws['H46'] = tdp
    ws['H46'].alignment = Alignment(horizontal='center')
    ws['I46'] = "=({}*{})".format(tpp,char_dict['lvl'])
    ws['I46'].alignment = Alignment(horizontal='center')
    ws['I46'].font = textFont
    ws['I46'].border = fullBorder

    ######## Skills sheet ########
    ws1['A1'] = "Skills"
    ws1['B1'] = "Cost"
    ws1['C1'] = "Stats"
    ws1.merge_cells('A1:A2')
    ws1.merge_cells('B1:B2')
    ws1.merge_cells('C1:C2')
    ws1.merge_cells('D1:H1')
    ws1.merge_cells('I1:M1')
    ws1['D1'] = "-- SKILL RANKS --"
    ws1['I1'] = "-- BONUSES --"
    ws1['D2'] = "Hobby"
    ws1['E2'] = "Adol"
    ws1['F2'] = "App"
    ws1['G2'] = "Ranks"
    ws1['H2'] = "Total"
    ws1['I2'] = "Skill"
    ws1['J2'] = "Stat"
    ws1['K2'] = "Lvl"
    ws1['L2'] = "Misc"
    ws1['M2'] = "Total"
    ws1['B1'].alignment = Alignment(horizontal='center')
    ws1['C1'].alignment = Alignment(horizontal='center')
    ws1['D1'].alignment = Alignment(horizontal='center')
    ws1['I1'].alignment = Alignment(horizontal='center')
    ws1['C2'].alignment = Alignment(horizontal='center')
    ws1['D2'].alignment = Alignment(horizontal='center')
    ws1['E2'].alignment = Alignment(horizontal='center')
    ws1['F2'].alignment = Alignment(horizontal='center')
    ws1['G2'].alignment = Alignment(horizontal='center')
    ws1['H2'].alignment = Alignment(horizontal='center')
    ws1['I2'].alignment = Alignment(horizontal='center')
    ws1['J2'].alignment = Alignment(horizontal='center')
    ws1['K2'].alignment = Alignment(horizontal='center')
    ws1['L2'].alignment = Alignment(horizontal='center')
    ws1['M2'].alignment = Alignment(horizontal='center')
    ws1['A1'].border = fullBorder
    ws1['A2'].border = fullBorder
    ws1['B1'].border = fullBorder
    ws1['B2'].border = fullBorder
    ws1['C1'].border = fullBorder
    ws1['C2'].border = fullBorder
    ws1['D1'].border = fullBorder
    ws1['E1'].border = tbBorder
    ws1['F1'].border = tbBorder
    ws1['G1'].border = tbBorder
    ws1['H1'].border = fullBorder
    ws1['I1'].border = fullBorder
    ws1['J1'].border = tbBorder
    ws1['K1'].border = tbBorder
    ws1['L1'].border = tbBorder
    ws1['L1'].border = tbBorder
    ws1['M1'].border = fullBorder
    ws1['C2'].border = fullBorder
    ws1['D2'].border = fullBorder
    ws1['E2'].border = fullBorder
    ws1['F2'].border = fullBorder
    ws1['G2'].border = fullBorder
    ws1['H2'].border = fullBorder
    ws1['I2'].border = fullBorder
    ws1['J2'].border = fullBorder
    ws1['K2'].border = fullBorder
    ws1['L2'].border = fullBorder
    ws1['M2'].border = fullBorder
    ws1['A1'].font = textCalBoldFont
    ws1['B1'].font = textCalBoldFont
    ws1['C1'].font = textCalBoldFont
    ws1['D1'].font = textCalBoldFont
    ws1['D2'].font = textCalBoldFont
    ws1['E2'].font = textCalBoldFont
    ws1['F2'].font = textCalBoldFont
    ws1['I1'].font = textCalBoldFont
    ws1['G2'].font = textCalBoldFont
    ws1['H2'].font = textCalBoldFont
    ws1['I2'].font = textCalBoldFont
    ws1['J2'].font = textCalBoldFont
    ws1['K2'].font = textCalBoldFont
    ws1['L2'].font = textCalBoldFont
    ws1['M2'].font = textCalBoldFont

    skill=[]
    ###### Test ######
    ws1.column_dimensions['A'].width = 28.0
    ws1.column_dimensions['B'].width = 5.5
    ws1.column_dimensions['C'].width = 10.0
    ws1.column_dimensions['D'].width = 6.5
    ws1.column_dimensions['E'].width = 6.5
    ws1.column_dimensions['F'].width = 6.5
    ws1.column_dimensions['G'].width = 5.5
    ws1.column_dimensions['H'].width = 5.0
    ws1.column_dimensions['I'].width = 5.0
    ws1.column_dimensions['J'].width = 5.0
    ws1.column_dimensions['K'].width = 5.0
    ws1.column_dimensions['L'].width = 5.0
    ws1.column_dimensions['M'].width = 7.0

    # Page Titles
    ws1.print_title_rows = '1:2'
    ws1.print_title_cols = 'A:M'

    skill_rank_total = 0
    for words in char_dict:
        if words.isdigit():
            if char_dict[words][5] > 0 or char_dict[words][6]>0 or char_dict[words][7]>0 or char_dict[words][8]>0:
                skill_rank_total = char_dict[words][5] + char_dict[words][6] + char_dict[words][7]+ char_dict[words][8]
                skill.append([char_dict[words][0],char_dict[words][3],char_dict[words][1],char_dict[words][5],char_dict[words][6],char_dict[words][7],char_dict[words][8],skill_rank_total,char_dict[words][10],char_dict[words][11],char_dict[words][13],char_dict[words][12],char_dict[words][14]])
                skill.sort()

    for row in skill:
        ws1.append(row)
    for row in ws1.iter_rows(min_row=4,min_col=1):
        pass

    # loop thru rows, then cells in the row to apply alignment and border
    rows,nrows=1,3
    while rows <= len(skill):
        for cell in ws1[nrows:nrows]:
            cell.border = fullBorder
            cell.alignment = Alignment(horizontal='center')
        rows+=1
        nrows+=1
    # Loop thru columns and alignment column 'A' (Skill Names)
    for col in ws.iter_cols(min_row=1, max_col=11):
        for cell in ws1['A']:
            cell.alignment = Alignment(horizontal='left')

    # Loop thru columns and set color fill on column 'M' (Skill Totals)
    for col in ws.iter_cols(min_row=1, max_col=11):
        for cell in ws1['M']:
            cell.fill = lblue
    # Save the file
    wb.save(charXlPath+"/"+char+"/"+charXlFile)
#export_to_excel()

def export_allskills_to_excel():
    p=charMenu.char_menu()
    menu_len=len(p)
    while True:
        s=int(raw_input("Select Character: "))
        if s >=1 and s<=menu_len:
            break
        else:
            print "Invalid Selection! Select a character from the list"

    # Open the file
    char_dict={}
    s-=1
    char = p[s]
    with open(cfgData.char_dir+"/"+char+"/"+char+".json","r") as cf:
        char_dict = json.load(cf)
    charXlFile2=char_dict['Fullname']+"_AllSkills.xlsx"

    print ""
    print "Creating excel file....."

    # Setup Excel file
    wb = Workbook()
    # grab the active worksheet
    ws = wb.active
    ws.title="Skills"
    ws1 = wb['Skills']
    ws.page_margins.bottom=0.5
    ws.page_margins.top=0.5
    ws1.page_margins.top=0.25
    ws1.page_margins.bottom=0.25
    ws1.page_margins.left=0.25
    ws1.page_margins.right=0.25

    # Define colors
    gray = PatternFill("solid",fgColor="969696")
    lblue = PatternFill("solid",fgColor="CCFFFF")

    # Define fonts
    textFont = Font(name='Arial',size=10)
    textBoldFont = Font(name='Arial',size=11, bold=True)
    textCalBoldFont = Font(name='Calibri',size=11, bold=True)
    headerFont = Font(name='Old English Text MT',size=14,bold=True)
    statFont = Font(name='Old English Text MT',size=18,bold=True)

    # Define Borders
    fullBorder = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
    tbBorder = Border(top=Side(style='thin'), bottom=Side(style='thin'))
    tblBorder = Border(top=Side(style='thin'), bottom=Side(style='thin'), left=Side(style='thin'))
    tbrBorder = Border(top=Side(style='thin'), bottom=Side(style='thin'), right=Side(style='thin'))
    rBorder = Border(right=Side(style='thin'))
    lBorder = Border(left=Side(style='thin'))
    tBorder = Border(top=Side(style='thin'))
    bBorder = Border(bottom=Side(style='thin'))
    tlrBorder = Border(top=Side(style='thin'), left=Side(style='thin'), right=Side(style='thin'))
    blrBorder = Border(left=Side(style='thin'), bottom=Side(style='thin'), right=Side(style='thin'))
    rlBorder = Border(left=Side(style='thin'), right=Side(style='thin'))
    rbBorder = Border(bottom=Side(style='thin'), right=Side(style='thin'))
    lbBorder = Border(bottom=Side(style='thin'), left=Side(style='thin'))
    tlBorder = Border(top=Side(style='thin'), left=Side(style='thin'))

    ws1['A1'] = "Skills"
    ws1['B1'] = "Cost"
    ws1['C1'] = "Stats"
    ws1.merge_cells('A1:A2')
    ws1.merge_cells('B1:B2')
    ws1.merge_cells('C1:C2')
    ws1.merge_cells('D1:H1')
    ws1.merge_cells('I1:M1')
    ws1['D1'] = "-- SKILL RANKS --"
    ws1['I1'] = "-- BONUSES --"
    ws1['D2'] = "Hobby"
    ws1['E2'] = "Adol"
    ws1['F2'] = "App"
    ws1['G2'] = "Ranks"
    ws1['H2'] = "Total"
    ws1['I2'] = "Skill"
    ws1['J2'] = "Stat"
    ws1['K2'] = "Lvl"
    ws1['L2'] = "Misc"
    ws1['M2'] = "Total"
    ws1['B1'].alignment = Alignment(horizontal='center')
    ws1['C1'].alignment = Alignment(horizontal='center')
    ws1['D1'].alignment = Alignment(horizontal='center')
    ws1['I1'].alignment = Alignment(horizontal='center')
    ws1['C2'].alignment = Alignment(horizontal='center')
    ws1['D2'].alignment = Alignment(horizontal='center')
    ws1['E2'].alignment = Alignment(horizontal='center')
    ws1['F2'].alignment = Alignment(horizontal='center')
    ws1['G2'].alignment = Alignment(horizontal='center')
    ws1['H2'].alignment = Alignment(horizontal='center')
    ws1['I2'].alignment = Alignment(horizontal='center')
    ws1['J2'].alignment = Alignment(horizontal='center')
    ws1['K2'].alignment = Alignment(horizontal='center')
    ws1['L2'].alignment = Alignment(horizontal='center')
    ws1['M2'].alignment = Alignment(horizontal='center')
    ws1['A1'].border = fullBorder
    ws1['A2'].border = fullBorder
    ws1['B1'].border = fullBorder
    ws1['B2'].border = fullBorder
    ws1['C1'].border = fullBorder
    ws1['C2'].border = fullBorder
    ws1['D1'].border = fullBorder
    ws1['E1'].border = tbBorder
    ws1['F1'].border = tbBorder
    ws1['G1'].border = tbBorder
    ws1['H1'].border = fullBorder
    ws1['I1'].border = fullBorder
    ws1['J1'].border = tbBorder
    ws1['K1'].border = tbBorder
    ws1['L1'].border = tbBorder
    ws1['L1'].border = tbBorder
    ws1['M1'].border = fullBorder
    ws1['C2'].border = fullBorder
    ws1['D2'].border = fullBorder
    ws1['E2'].border = fullBorder
    ws1['F2'].border = fullBorder
    ws1['G2'].border = fullBorder
    ws1['H2'].border = fullBorder
    ws1['I2'].border = fullBorder
    ws1['J2'].border = fullBorder
    ws1['K2'].border = fullBorder
    ws1['L2'].border = fullBorder
    ws1['M2'].border = fullBorder
    ws1['A1'].font = textCalBoldFont
    ws1['B1'].font = textCalBoldFont
    ws1['C1'].font = textCalBoldFont
    ws1['D1'].font = textCalBoldFont
    ws1['D2'].font = textCalBoldFont
    ws1['E2'].font = textCalBoldFont
    ws1['F2'].font = textCalBoldFont
    ws1['I1'].font = textCalBoldFont
    ws1['G2'].font = textCalBoldFont
    ws1['H2'].font = textCalBoldFont
    ws1['I2'].font = textCalBoldFont
    ws1['J2'].font = textCalBoldFont
    ws1['K2'].font = textCalBoldFont
    ws1['L2'].font = textCalBoldFont
    ws1['M2'].font = textCalBoldFont

    skill=[]
    ###### Test ######
    ws1.column_dimensions['A'].width = 28.0
    ws1.column_dimensions['B'].width = 5.5
    ws1.column_dimensions['C'].width = 10.0
    ws1.column_dimensions['D'].width = 6.5
    ws1.column_dimensions['E'].width = 6.5
    ws1.column_dimensions['F'].width = 6.5
    ws1.column_dimensions['G'].width = 5.5
    ws1.column_dimensions['H'].width = 5.0
    ws1.column_dimensions['I'].width = 5.0
    ws1.column_dimensions['J'].width = 5.0
    ws1.column_dimensions['K'].width = 5.0
    ws1.column_dimensions['L'].width = 5.0
    ws1.column_dimensions['M'].width = 7.0

    # Page Titles
    ws1.print_title_rows = '1:2'
    ws1.print_title_cols = 'A:M'

    with open(cfgData.char_dir+"/"+char+"/"+char+".json","r") as cf:
        char_dict = json.load(cf)
    with open(cfgData.char_dir+"/"+char+"/"+char+"_Skills.json","r") as sf:
        char_skill = json.load(sf)
    charXlFile2=char_dict['Fullname']+"_AllSkills.xlsx"

    skill_rank_total = 0
    for words in char_skill:
        skill_rank_total = char_skill[words][5] + char_skill[words][6] + char_skill[words][7]+ char_skill[words][8]
        skill.append([char_skill[words][0],char_skill[words][3],char_skill[words][1],char_skill[words][5],char_skill[words][6],char_skill[words][7],char_skill[words][8],skill_rank_total,char_skill[words][10],char_skill[words][11],char_skill[words][13],char_skill[words][12],char_skill[words][14]])
        skill.sort()

    for row in skill:
        ws1.append(row)
    for row in ws1.iter_rows(min_row=4,min_col=1):
        pass

    # loop thru rows, then cells in the row to apply alignment and border
    rows,nrows=1,3
    while rows <= len(skill):
        for cell in ws1[nrows:nrows]:
            cell.border = fullBorder
            cell.alignment = Alignment(horizontal='center')
        rows+=1
        nrows+=1
    # Loop thru columns and alignment column 'A' (Skill Names)
    for col in ws.iter_cols(min_row=1, max_col=11):
        for cell in ws1['A']:
            cell.alignment = Alignment(horizontal='left')

    # Loop thru columns and set color fill on column 'M' (Skill Totals)
    for col in ws.iter_cols(min_row=1, max_col=11):
        for cell in ws1['M']:
            cell.fill = lblue
    # Save the file
    wb.save(cfgData.char_dir+"/"+char+"/"+charXlFile2)
